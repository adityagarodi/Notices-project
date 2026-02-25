from datetime import datetime, timezone, timedelta
from flask import Flask, request, jsonify, send_from_directory, url_for
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.exc import IntegrityError
from sqlalchemy import text
from werkzeug.security import generate_password_hash, check_password_hash
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import os
from datetime import date as date_cls
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge

# -----------------------------------------------------------------------------
# App and Database Setup
# -----------------------------------------------------------------------------
app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, 'notices.db')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-change-me')
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024  # 20 MB total request cap (mobile friendly)

db = SQLAlchemy(app)


# -----------------------------------------------------------------------------
# Models
# -----------------------------------------------------------------------------
class Notice(db.Model):
    __tablename__ = 'notices'

    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(255), nullable=False)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=lambda: datetime.now(timezone.utc))
    notice_date = db.Column(db.Date, nullable=True)
    image_filename = db.Column(db.String(255), nullable=True)

    def to_dict(self):
        data = {
            'id': self.id,
            'title': self.title,
            'content': self.content,
            'created_at': self.created_at.isoformat(),
        }
        if self.image_filename:
            # Build absolute URL to serve image
            try:
                img_url = url_for('serve_upload', filename=self.image_filename, _external=True)
            except RuntimeError:
                # Fallback if app context missing; construct manually
                img_url = f"/uploads/{self.image_filename}"
            data['image_url'] = img_url
        return data


class User(db.Model):
    __tablename__ = 'users'

    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=lambda: datetime.now(timezone.utc))

    def to_dict(self):
        return {
            'id': self.id,
            'username': self.username,
            'created_at': self.created_at.isoformat(),
        }


# Create DB tables if not exist
with app.app_context():
    db.create_all()

    # Ensure schema has new columns when upgrading existing DB
    try:
        cols = {row[1] for row in db.session.execute(text('PRAGMA table_info(notices)')).fetchall()}
        if 'notice_date' not in cols:
            db.session.execute(text('ALTER TABLE notices ADD COLUMN notice_date DATE'))
        if 'image_filename' not in cols:
            db.session.execute(text('ALTER TABLE notices ADD COLUMN image_filename VARCHAR(255)'))
        # Ensure User table has advanced auth columns
        ucols = {row[1] for row in db.session.execute(text('PRAGMA table_info(users)')).fetchall()}
        if 'failed_attempts' not in ucols:
            db.session.execute(text('ALTER TABLE users ADD COLUMN failed_attempts INTEGER NOT NULL DEFAULT 0'))
        if 'locked_until' not in ucols:
            db.session.execute(text('ALTER TABLE users ADD COLUMN locked_until DATETIME NULL'))
        db.session.commit()
    except Exception:
        db.session.rollback()


# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------

def json_error(message, status=400):
    return jsonify({'error': message}), status


def get_json():
    data = request.get_json(silent=True)
    if data is None:
        raise ValueError('Request body must be JSON')
    return data


def validate_notice_payload(data, partial=False):
    """
    Validate incoming JSON payload.
    If partial=True, allow missing fields (for PUT/PATCH-like update).
    """
    if not isinstance(data, dict):
        raise ValueError('JSON body must be an object')

    if not partial or 'title' in data:
        title = data.get('title')
        if not isinstance(title, str) or not title.strip():
            raise ValueError('Field "title" is required and must be a non-empty string')

    if not partial or 'content' in data:
        content = data.get('content')
        if not isinstance(content, str) or not content.strip():
            raise ValueError('Field "content" is required and must be a non-empty string')

ALLOWED_IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.heic', '.heif'}
ALLOWED_MIME_TYPES = {'image/png', 'image/jpeg', 'image/gif', 'image/heic', 'image/heif'}
MAX_IMAGE_BYTES = 10 * 1024 * 1024  # 10 MB per image (mobile friendly)

def validate_image(file):
    if not file or file.filename == '':
        return None
    filename = secure_filename(file.filename)
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_IMAGE_EXTENSIONS:
        raise ValueError('Unsupported image type. Allowed: PNG, JPG, JPEG, GIF, HEIC, HEIF')
    # Validate mimetype if provided
    if file.mimetype and file.mimetype not in ALLOWED_MIME_TYPES:
        raise ValueError('Invalid image content type')
    # Validate size
    try:
        pos = file.stream.tell()
        file.stream.seek(0, os.SEEK_END)
        size = file.stream.tell()
        file.stream.seek(pos)
        if size and size > MAX_IMAGE_BYTES:
            raise ValueError('Image too large (max 10 MB)')
    except Exception:
        # If we fail to detect, rely on MAX_CONTENT_LENGTH to guard huge uploads
        pass
    # Return the sanitized filename for downstream use
    return filename


def ensure_auth_workbook():
    """Ensure auth Excel workbook exists with sheets and headers."""
    os.makedirs(os.path.dirname(AUTH_XLSX), exist_ok=True)
    if not os.path.exists(AUTH_XLSX):
        wb = Workbook()
        # Sheet for user credentials (hashed)
        ws_users = wb.active
        ws_users.title = 'users'
        ws_users.append(['id', 'username', 'password_hash', 'created_at_iso'])
        for cell in ws_users[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical='center')
        ws_users.column_dimensions['A'].width = 8
        ws_users.column_dimensions['B'].width = 24
        ws_users.column_dimensions['C'].width = 64
        ws_users.column_dimensions['D'].width = 26
        # Sheet for successful login attempts
        ws_logins = wb.create_sheet('logins')
        ws_logins.append(['timestamp_utc', 'username', 'ip', 'user_agent'])
        for cell in ws_logins[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical='center')
        ws_logins.column_dimensions['A'].width = 26
        ws_logins.column_dimensions['B'].width = 24
        ws_logins.column_dimensions['C'].width = 20
        ws_logins.column_dimensions['D'].width = 50
        wb.save(AUTH_XLSX)


def log_user_credentials(user: 'User'):
    """Append a user's hashed credentials to auth.xlsx (no plain passwords)."""
    try:
        ensure_auth_workbook()
        wb = load_workbook(AUTH_XLSX)
        ws = wb['users'] if 'users' in wb.sheetnames else wb.active
        ws.append([
            getattr(user, 'id', None),
            getattr(user, 'username', ''),
            getattr(user, 'password_hash', ''),
            getattr(user, 'created_at', datetime.now(timezone.utc)).isoformat(),
        ])
        wb.save(AUTH_XLSX)
    except Exception:
        # Don't fail API if logging fails
        pass


def log_login_success(username: str, ip: str, user_agent: str):
    """Append a successful login record to auth.xlsx."""
    try:
        ensure_auth_workbook()
        wb = load_workbook(AUTH_XLSX)
        ws = wb['logins'] if 'logins' in wb.sheetnames else wb.active
        ws.append([
            datetime.now(timezone.utc).isoformat(),
            username,
            ip,
            user_agent[:255] if user_agent else '',
        ])
        wb.save(AUTH_XLSX)
    except Exception:
        pass

def log_login_attempt(username: str, success: bool, ip: str, user_agent: str, reason: str = ''):
    """Log any login attempt result into auth.xlsx/attempts."""
    try:
        ensure_auth_workbook()
        wb = load_workbook(AUTH_XLSX)
        ws = wb['attempts'] if 'attempts' in wb.sheetnames else wb.create_sheet('attempts')
        if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
            # add header if empty
            ws.append(['timestamp_utc', 'username', 'success', 'ip', 'user_agent', 'reason'])
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(vertical='center')
            ws.column_dimensions['A'].width = 26
            ws.column_dimensions['B'].width = 24
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 50
            ws.column_dimensions['F'].width = 24
        ws.append([
            datetime.now(timezone.utc).isoformat(),
            username or '',
            bool(success),
            ip or '',
            (user_agent or '')[:255],
            reason or '',
        ])
        wb.save(AUTH_XLSX)
    except Exception:
        pass

# History workbook: login history sheet with formatting
def ensure_login_history_sheet(wb: Workbook):
    if 'LoginHistory' in wb.sheetnames:
        return wb['LoginHistory']
    ws = wb.create_sheet('LoginHistory')
    ws.append(['timestamp_utc', 'username', 'result', 'ip', 'user_agent', 'reason'])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical='center')
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 24
    return ws

def log_login_history(username: str, success: bool, ip: str, user_agent: str, reason: str = ''):
    try:
        ensure_history_workbook()
        wb = load_workbook(HISTORY_XLSX)
        ws = ensure_login_history_sheet(wb)
        ws.append([
            datetime.now(timezone.utc).isoformat(),
            username or '',
            'success' if success else 'failure',
            ip or '',
            (user_agent or '')[:255],
            reason or '',
        ])
        wb.save(HISTORY_XLSX)
    except Exception:
        pass

def save_image(file):
    if not file or file.filename == '':
        return None
    filename = validate_image(file)
    # Prefix with timestamp to avoid collisions
    ts = datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S%f')
    name, ext = os.path.splitext(filename)
    stored = f"{name}_{ts}{ext}"
    path = os.path.join(UPLOAD_DIR, stored)
    file.save(path)
    return stored


# -----------------------------------------------------------------------------
# Excel logging for notice history and auth
# -----------------------------------------------------------------------------

# Store history.xlsx in the project root (one level above backend) to keep it portable
# Avoid hardcoded absolute paths so the app runs on any machine
PROJECT_ROOT = os.path.abspath(os.path.join(BASE_DIR, os.pardir))
HISTORY_XLSX = os.path.join(PROJECT_ROOT, 'history.xlsx')
AUTH_XLSX = os.path.join(PROJECT_ROOT, 'auth.xlsx')
AUTH_LOGGING_ENABLED = False  # Disable storing any auth details to disk

# Simple Excel-based auth (two columns: Username, Password)
def ensure_simple_auth_workbook():
    """Ensure a very simple auth.xlsx exists with a single sheet and headers.
    Sheet name: 'auth', Columns: ['Username', 'Password']
    """
    os.makedirs(os.path.dirname(AUTH_XLSX), exist_ok=True)
    if not os.path.exists(AUTH_XLSX):
        wb = Workbook()
        ws = wb.active
        ws.title = 'auth'
        ws.append(['Username', 'Password'])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical='center')
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 24
        wb.save(AUTH_XLSX)

def write_simple_auth(username: str, password: str):
    """Overwrite auth.xlsx with the provided single credential row."""
    ensure_simple_auth_workbook()
    wb = Workbook()
    ws = wb.active
    ws.title = 'auth'
    ws.append(['Username', 'Password'])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical='center')
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 24
    ws.append([username, password])
    wb.save(AUTH_XLSX)

def read_simple_auth():
    """Read the single credential from auth.xlsx. Returns (username, password) or (None, None)."""
    try:
        ensure_simple_auth_workbook()
        wb = load_workbook(AUTH_XLSX)
        ws = wb['auth'] if 'auth' in wb.sheetnames else wb.active
        # Expect header in row 1, data in row 2
        if ws.max_row >= 2:
            u = (ws.cell(row=2, column=1).value or '').strip()
            p = ws.cell(row=2, column=2).value or ''
            return u, p
        return None, None
    except Exception:
        return None, None


def ensure_history_workbook():
    """Ensure history Excel workbook exists with headers."""
    # Ensure target directory exists
    os.makedirs(os.path.dirname(HISTORY_XLSX), exist_ok=True)
    if not os.path.exists(HISTORY_XLSX):
        wb = Workbook()
        ws = wb.active
        ws.title = 'History'
        ws.append(['timestamp_utc', 'action', 'notice_id', 'title', 'content'])
        wb.save(HISTORY_XLSX)


def log_notice_event(action: str, notice: 'Notice'):
    try:
        ensure_history_workbook()
        wb = load_workbook(HISTORY_XLSX)
        ws = wb.active
        ws.append([
            datetime.now(timezone.utc).isoformat(),
            action,
            getattr(notice, 'id', None),
            getattr(notice, 'title', ''),
            getattr(notice, 'content', ''),
        ])
        wb.save(HISTORY_XLSX)
    except Exception as e:
        # Do not break the API if history logging fails
        try:
            app.logger.exception('History logging failed')
        except Exception:
            pass


# -----------------------------------------------------------------------------
# Auth helpers and decorators
# -----------------------------------------------------------------------------

def serializer():
    return URLSafeTimedSerializer(app.config['SECRET_KEY'], salt='auth-token')


def create_token(user_id: int) -> str:
    return serializer().dumps({'uid': user_id})


def verify_token(token: str, max_age_seconds: int = 60 * 60 * 24 * 7):  # 7 days
    try:
        data = serializer().loads(token, max_age=max_age_seconds)
        return data.get('uid')
    except (BadSignature, SignatureExpired):
        return None


def require_auth(func):
    from functools import wraps

    @wraps(func)
    def wrapper(*args, **kwargs):
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            return json_error('Unauthorized', 401)
        token = auth_header.split(' ', 1)[1]
        uid = verify_token(token)
        if not uid:
            return json_error('Unauthorized', 401)
        # Optionally attach user to request context if needed
        return func(*args, **kwargs)

    return wrapper


# -----------------------------------------------------------------------------
# Rate limiting (simple in-memory per-IP attempts for login)
# -----------------------------------------------------------------------------
RATE_LIMIT_WINDOW_SEC = 60  # 1 minute window
RATE_LIMIT_MAX_ATTEMPTS = 15  # max attempts per IP per window
_rate_limit_cache = {}

def is_rate_limited(ip: str):
    now = datetime.now(timezone.utc)
    window_start = now - timedelta(seconds=RATE_LIMIT_WINDOW_SEC)
    entry = _rate_limit_cache.get(ip)
    if not entry:
        _rate_limit_cache[ip] = {'timestamps': [now]}
        return False
    # prune old
    entry['timestamps'] = [t for t in entry['timestamps'] if t >= window_start]
    entry['timestamps'].append(now)
    _rate_limit_cache[ip] = entry
    return len(entry['timestamps']) > RATE_LIMIT_MAX_ATTEMPTS

# Account lock policy
LOCK_THRESHOLD = 5  # failed attempts before lock
LOCK_DURATION_MIN = 15

def password_is_strong(pw: str) -> bool:
    """Relaxed policy to allow simple passwords like '123456'."""
    return isinstance(pw, str) and len(pw) >= 1

# -----------------------------------------------------------------------------
# Routes (CRUD)
# -----------------------------------------------------------------------------
@app.route('/api/notices', methods=['POST'])
@require_auth
def create_notice():
    try:
        if request.content_type and request.content_type.startswith('multipart/form-data'):
            title = (request.form.get('title') or '').strip()
            content = (request.form.get('content') or '').strip()
            if not title:
                return json_error('Field "title" is required', 400)
            if not content:
                return json_error('Field "content" is required', 400)
            image_file = request.files.get('image')
            # Support selecting an existing image from media library
            selected_image_filename = (request.form.get('image_filename') or '').strip()
            image_filename = None
            if image_file and image_file.filename:
                image_filename = save_image(image_file)
            elif selected_image_filename:
                # Only accept if file exists in uploads and has allowed extension
                candidate_path = os.path.join(UPLOAD_DIR, os.path.basename(selected_image_filename))
                ext = os.path.splitext(candidate_path)[1].lower()
                if os.path.isfile(candidate_path) and ext in ALLOWED_IMAGE_EXTENSIONS:
                    image_filename = os.path.basename(selected_image_filename)
            notice = Notice(title=title, content=content, image_filename=image_filename)
        else:
            data = get_json()
            validate_notice_payload(data)
            notice = Notice(
                title=data['title'].strip(),
                content=data['content'].strip(),
            )
        db.session.add(notice)
        db.session.commit()
        log_notice_event('CREATE', notice)
        return jsonify(notice.to_dict()), 201
    except ValueError as ve:
        return json_error(str(ve), 400)
    except IntegrityError as ie:
        db.session.rollback()
        try:
            app.logger.exception('DB integrity error on create_notice')
        except Exception:
            pass
        return json_error('Database integrity error', 400)
    except Exception as e:
        try:
            app.logger.exception('create_notice failed')
        except Exception:
            pass
        return json_error('Internal server error', 500)


@app.route('/api/notices', methods=['GET'])
def list_notices():
    notices = Notice.query.order_by(Notice.created_at.desc()).all()
    return jsonify([n.to_dict() for n in notices])


@app.route('/api/notices/<int:notice_id>', methods=['GET'])
def get_notice(notice_id: int):
    notice = Notice.query.get(notice_id)
    if not notice:
        return json_error('Notice not found', 404)
    return jsonify(notice.to_dict())


@app.route('/api/notices/<int:notice_id>', methods=['PUT'])
@require_auth
def update_notice(notice_id: int):
    notice = Notice.query.get(notice_id)
    if not notice:
        return json_error('Notice not found', 404)

    try:
        if request.content_type and request.content_type.startswith('multipart/form-data'):
            title = request.form.get('title')
            content = request.form.get('content')
            if title is not None:
                if not title.strip():
                    return json_error('Field "title" must be non-empty', 400)
                notice.title = title.strip()
            if content is not None:
                if not content.strip():
                    return json_error('Field "content" must be non-empty', 400)
                notice.content = content.strip()
            image_file = request.files.get('image')
            selected_image_filename = (request.form.get('image_filename') or '').strip()
            if image_file and image_file.filename:
                stored = save_image(image_file)
                notice.image_filename = stored
            elif selected_image_filename:
                candidate_path = os.path.join(UPLOAD_DIR, os.path.basename(selected_image_filename))
                ext = os.path.splitext(candidate_path)[1].lower()
                if os.path.isfile(candidate_path) and ext in ALLOWED_IMAGE_EXTENSIONS:
                    notice.image_filename = os.path.basename(selected_image_filename)
        else:
            data = get_json()
            # full update; allow partial fields but validate if present
            validate_notice_payload(data, partial=True)
            if 'title' in data:
                notice.title = data['title'].strip()
            if 'content' in data:
                notice.content = data['content'].strip()

        db.session.commit()
        log_notice_event('UPDATE', notice)
        return jsonify(notice.to_dict())
    except ValueError as ve:
        return json_error(str(ve), 400)
    except IntegrityError as ie:
        db.session.rollback()
        try:
            app.logger.exception('DB integrity error on update_notice')
        except Exception:
            pass
        return json_error('Database integrity error', 400)
    except Exception as e:
        try:
            app.logger.exception('update_notice failed')
        except Exception:
            pass
        return json_error('Internal server error', 500)


# -----------------------------------------------------------------------------
# Media Library
# -----------------------------------------------------------------------------
@app.route('/api/media', methods=['GET'])
def list_media():
    """List uploaded media files in the uploads directory."""
    try:
        files = []
        for fname in os.listdir(UPLOAD_DIR):
            path = os.path.join(UPLOAD_DIR, fname)
            if not os.path.isfile(path):
                continue
            ext = os.path.splitext(fname)[1].lower()
            if ext not in ALLOWED_IMAGE_EXTENSIONS:
                continue
            stat = os.stat(path)
            files.append({
                'filename': fname,
                'url': url_for('serve_upload', filename=fname, _external=True),
                'size_bytes': stat.st_size,
                'modified_at': datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
            })
        # Newest first
        files.sort(key=lambda f: f['modified_at'], reverse=True)
        return jsonify(files)
    except Exception:
        return jsonify([])


@app.route('/api/notices/<int:notice_id>', methods=['DELETE'])
@require_auth
def delete_notice(notice_id: int):
    notice = Notice.query.get(notice_id)
    if not notice:
        return json_error('Notice not found', 404)

    try:
        db.session.delete(notice)
        db.session.commit()
        log_notice_event('DELETE', notice)
        return jsonify({'message': 'Deleted'}), 200
    except Exception as e:
        db.session.rollback()
        try:
            app.logger.exception('delete_notice failed')
        except Exception:
            pass
        return json_error('Internal server error', 500)


# -----------------------------------------------------------------------------
# Health Check
# -----------------------------------------------------------------------------
@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'}), 200


# -----------------------------------------------------------------------------
# Auth Routes
# -----------------------------------------------------------------------------
@app.route('/api/auth/register', methods=['POST'])
def register():
    """Register a new account by storing credentials in auth.xlsx (Username, Password).
    Overwrites any previous data to keep only the latest account.
    """
    data = request.get_json(silent=True)
    if not data or not isinstance(data, dict):
        return json_error('Request body must be JSON', 400)

    username = (data.get('username') or '').strip()
    password = data.get('password') or ''

    if not username:
        return json_error('Username is required', 400)
    # No password validation per request; allow simple values like '123456'
    if not isinstance(password, str) or len(password) < 1:
        return json_error('Password is required', 400)

    # Write credentials to Excel, overwriting previous data
    try:
        write_simple_auth(username, password)
    except Exception:
        return json_error('Failed to write credentials', 500)

    # Issue a token (not tied to DB). Admin routes just require a signed token.
    token = create_token(1)
    return jsonify({
        'message': 'Account created successfully',
        'user': {'username': username},
        'token': token,
    }), 201


@app.route('/api/auth/login', methods=['POST'])
def login():
    """Validate credentials by reading auth.xlsx (Username, Password)."""
    data = request.get_json(silent=True)
    if not data or not isinstance(data, dict):
        return json_error('Request body must be JSON', 400)

    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    if not username or not password:
        return json_error('Username and password are required', 400)

    # Read stored credentials from Excel
    stored_user, stored_pass = read_simple_auth()
    if stored_user is None:
        return json_error('No account found. Please create an account first.', 404)

    if username == stored_user and password == stored_pass:
        token = create_token(1)
        return jsonify({
            'message': 'Login successful',
            'user': {'username': username},
            'token': token,
        }), 200
    else:
        return json_error('Invalid username or password', 401)


# Serve uploaded images
@app.route('/uploads/<path:filename>')
def serve_upload(filename):
    return send_from_directory(UPLOAD_DIR, filename)


# Global error handlers
@app.errorhandler(RequestEntityTooLarge)
def handle_request_entity_too_large(e):
    try:
        app.logger.warning('Payload too large')
    except Exception:
        pass
    return jsonify({'error': 'Upload too large. Max total request size is 20 MB and image max is 10 MB.'}), 413


if __name__ == '__main__':
    # For local development; use a proper WSGI server in production
    # Bind to 0.0.0.0 so other devices on your LAN (e.g., your phone) can access it
    app.run(host='0.0.0.0', port=5000, debug=True)
